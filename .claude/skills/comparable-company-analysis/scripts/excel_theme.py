"""
Build themed Excel for comparable-company analysis.
"""

import argparse
import asyncio
import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from get_data import fetch_comparable_company_data


DEFAULT_OUTPUT_DIR = Path.cwd() / "miaoxiang" / "comparable_company_analysis"

SECTION_FINANCE_HEADERS = [
    "股票代码",
    "营业总收入",
    "营业利润",
    "净利润",
    "流动资产",
    "资产总计",
    "流动负债总计",
    "负债合计",
    "所有者权益合计",
    "经营活动产生的现金流量净额",
    "投资活动产生的现金流量净额",
    "筹资活动产生的现金流量净额",
]

SECTION_VALUATION_HEADERS = [
    "股票代码",
    "最新收盘价(元)",
    "涨跌幅(%)",
    "总市值",
    "流通市值",
    "每股收益，TTM",
    "每股收益，26E",
    "市盈率PE，TTM",
    "市盈率PE，26E",
    "市净率PB(MRQ)",
    "市现率PCF(TTM)",
    "市销率PS(TTM)",
]

HIGHLIGHT_ROW_NAMES = {
    "行业均值",
    "行业中值",
    "最大值",
    "中位数",
    "最小值",
    "VS中位数(%,目标公司)",
    "Z-Score(目标公司)",
}



def _safe_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _coerce_numeric(value: Any) -> Any:
    """
    Convert numeric-like text to numbers for Excel cells.
    Keep non-numeric tokens (e.g. '-', stock codes) unchanged.
    """
    if isinstance(value, (int, float)):
        return value

    text = _safe_text(value)
    if not text or text == "-":
        return value

    normalized = text.replace(",", "").strip()
    if normalized.endswith("%"):
        normalized = normalized[:-1].strip()

    # Skip values that obviously contain letters or Chinese characters.
    if re.search(r"[A-Za-z\u4e00-\u9fff]", normalized):
        return value

    try:
        num = float(normalized)
    except Exception:
        return value

    if num.is_integer():
        return int(num)
    return num


def _thousand_number_format_from_raw(value: Any) -> Optional[str]:
    """
    Build Excel number format with thousand separators and
    default two decimal places.
    """
    text = _safe_text(value)
    if not text or text == "-":
        return None

    normalized = text.replace(",", "").strip()
    if normalized.endswith("%"):
        normalized = normalized[:-1].strip()
    if not normalized:
        return None

    return "#,##0.00"


def _normalize_rows(table: Dict[str, Any], expected_cols: int) -> List[List[str]]:
    rows: List[List[str]] = []
    if not isinstance(table, dict):
        return rows

    for row_name, row_values in table.items():
        if row_name == "headName":
            continue
        values = row_values if isinstance(row_values, list) else []
        normalized = [_safe_text(v, "-") for v in values[:expected_cols]]
        if len(normalized) < expected_cols:
            normalized.extend(["-"] * (expected_cols - len(normalized)))
        rows.append([_safe_text(row_name, "-")] + normalized)
    return rows


def _set_row_fill(ws, row_idx: int, start_col: int, end_col: int, fill: PatternFill) -> None:
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _set_row_font(ws, row_idx: int, start_col: int, end_col: int, font: Font) -> None:
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row=row_idx, column=col_idx).font = font


def _set_row_alignment(ws, row_idx: int, start_col: int, end_col: int, alignment: Alignment) -> None:
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row=row_idx, column=col_idx).alignment = alignment


def _set_table_border(ws, row_start: int, row_end: int, col_start: int, col_end: int) -> None:
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = border


def _write_section_table(
    ws,
    start_row: int,
    section_title: str,
    headers: List[str],
    rows: List[List[str]],
    total_cols: int,
) -> int:
    section_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_fill = PatternFill(fill_type="solid", start_color="DCE6F1", end_color="DCE6F1")
    white_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    highlight_fill = PatternFill(fill_type="solid", start_color="E6E6E6", end_color="E6E6E6")
    first_data_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")

    white_font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F1F1F")
    body_font = Font(name="Microsoft YaHei", size=10, color="1F1F1F")

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    ws.cell(row=start_row, column=1, value=section_title)
    _set_row_fill(ws, start_row, 1, total_cols, section_fill)
    _set_row_font(ws, start_row, 1, total_cols, white_font)
    _set_row_alignment(ws, start_row, 1, total_cols, Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[start_row].height = 24

    header_row = start_row + 1
    table_headers = ["公司名称"] + headers
    for idx, name in enumerate(table_headers, start=1):
        ws.cell(row=header_row, column=idx, value=name)
    _set_row_fill(ws, header_row, 1, total_cols, header_fill)
    _set_row_font(ws, header_row, 1, total_cols, header_font)
    _set_row_alignment(
        ws,
        header_row,
        1,
        total_cols,
        Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    ws.row_dimensions[header_row].height = 34

    current_row = header_row + 1
    for row_idx, values in enumerate(rows):
        for col_idx, value in enumerate(values, start=1):
            cell_value = value
            # A列=公司名称，B列=股票代码；从C列开始尝试转数字。
            if col_idx >= 3:
                cell_value = _coerce_numeric(value)
            cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
            if col_idx >= 3 and isinstance(cell_value, (int, float)):
                num_fmt = _thousand_number_format_from_raw(value)
                if num_fmt:
                    cell.number_format = num_fmt
        _set_row_fill(ws, current_row, 1, total_cols, white_fill)
        row_name = values[0] if values else ""
        if row_name in HIGHLIGHT_ROW_NAMES:
            ws.cell(row=current_row, column=1).fill = highlight_fill
        if row_idx == 0:
            for col_idx in range(1, total_cols + 1):
                ws.cell(row=current_row, column=col_idx).fill = first_data_fill
        _set_row_font(ws, current_row, 1, total_cols, body_font)
        _set_row_alignment(
            ws,
            current_row,
            1,
            total_cols,
            Alignment(horizontal="center", vertical="center", wrap_text=True),
        )
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    data_end = current_row - 1 if current_row > header_row else header_row
    _set_table_border(ws, header_row, data_end, 1, total_cols)
    return current_row


def build_themed_excel(
    payload: Dict[str, Any],
    output_dir: Optional[Path] = None,
    output_filename: Optional[str] = None,
) -> Dict[str, Any]:
    header = payload.get("header") if isinstance(payload, dict) else {}
    section_finance = payload.get("section_finance") if isinstance(payload, dict) else {}
    section_valuation = payload.get("section_valuation") if isinstance(payload, dict) else {}

    title = _safe_text(header.get("title"), "可比公司分析")
    input_title = _safe_text(header.get("inputTitle"), "")
    frontend_title = _safe_text(header.get("frontendTitle"), "")

    finance_table = section_finance.get("table") if isinstance(section_finance, dict) else {}
    valuation_table = section_valuation.get("table") if isinstance(section_valuation, dict) else {}

    finance_rows = _normalize_rows(finance_table, len(SECTION_FINANCE_HEADERS))
    valuation_rows = _normalize_rows(valuation_table, len(SECTION_VALUATION_HEADERS))

    workbook = Workbook()
    ws = workbook.active
    ws.title = "可比公司分析"

    total_cols = 13  # 公司名称 + 12 指标
    top_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    top_font_title = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    top_font_sub = Font(name="Microsoft YaHei", size=10, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)

    ws.cell(row=1, column=1, value=title)
    ws.cell(row=2, column=1, value=input_title)
    ws.cell(row=3, column=1, value=frontend_title)

    for top_row in (1, 2, 3):
        _set_row_fill(ws, top_row, 1, total_cols, top_fill)
        _set_row_alignment(ws, top_row, 1, total_cols, center)
    _set_row_font(ws, 1, 1, total_cols, top_font_title)
    _set_row_font(ws, 2, 1, total_cols, top_font_sub)
    _set_row_font(ws, 3, 1, total_cols, top_font_sub)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20

    current_row = 5
    current_row = _write_section_table(
        ws=ws,
        start_row=current_row,
        section_title="第一节：经营统计与财务指标",
        headers=SECTION_FINANCE_HEADERS,
        rows=finance_rows,
        total_cols=total_cols,
    )

    current_row += 3

    _write_section_table(
        ws=ws,
        start_row=current_row,
        section_title="第二节：估值情况",
        headers=SECTION_VALUATION_HEADERS,
        rows=valuation_rows,
        total_cols=total_cols,
    )

    col_widths = [18, 13, 17, 17, 17, 17, 18, 18, 18, 17, 17, 17, 17]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    out_dir = Path(output_dir or DEFAULT_OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    file_name = output_filename or "comparable_company_analysis_{0}.xlsx".format(uuid.uuid4().hex[:8])
    output_path = out_dir / file_name
    workbook.save(str(output_path))

    return {"output_path": str(output_path), "title": title, "row_count_finance": len(finance_rows), "row_count_valuation": len(valuation_rows)}


async def generate_excel_from_query(
    query: str,
    output_dir: Optional[Path] = None,
    output_filename: Optional[str] = None,
) -> Dict[str, Any]:
    data = await fetch_comparable_company_data(query)
    if "error" in data:
        return {"error": data["error"], "raw": data.get("raw")}
    return build_themed_excel(payload=data, output_dir=output_dir, output_filename=output_filename)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Create themed comparable-company Excel report.")
    parser.add_argument("--query", type=str, help="Natural language company question.")
    parser.add_argument("--output", type=str, default="", help="Optional output xlsx file path.")
    return parser


def run_cli() -> None:
    parser = _build_arg_parser()
    args = parser.parse_args()
    query = _safe_text(args.query)
    if not query:
        import sys

        query = _safe_text(sys.stdin.read())
    if not query:
        parser.print_help()
        raise SystemExit(1)

    output_path_raw = _safe_text(args.output)
    output_dir = None
    output_filename = None
    if output_path_raw:
        output_path = Path(output_path_raw)
        output_dir = output_path.parent
        output_filename = output_path.name

    async def _main() -> None:
        result = await generate_excel_from_query(
            query=query,
            output_dir=output_dir,
            output_filename=output_filename,
        )
        if "error" in result:
            print("Error: {0}".format(result["error"]))
            raise SystemExit(2)
        print("Saved: {0}".format(result.get("output_path", "")))

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_main())
    finally:
        loop.close()


if __name__ == "__main__":
    run_cli()
